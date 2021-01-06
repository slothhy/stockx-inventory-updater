import requests
import os
from openpyxl import load_workbook
from forex_python.converter import CurrencyRates
from datetime import datetime
import json
import time
import collections

__location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))

sales_db = collections.defaultdict(dict)
productname_db = collections.defaultdict()

def load_config():
    with open(os.path.join(__location__, "config.json")) as json_file:
        data = json.load(json_file)
        return data

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS  # pylint: disable=no-member
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)

def main():
    config = load_config()
   
    if config["currency"] == "USD":
        rate = 1 
    else:
        c = CurrencyRates()
        rate = c.get_rate('USD', config["currency"])

    wb = load_workbook(resource_path("./stock_book.xlsx"))
    ws = wb['Sheet1']

    session = requests.Session()
    session.headers.update({
        'content-type': 'application/x-www-form-urlencoded',
        'user-agent': config["user-agent"],
        'accept': '*/*',
        'accept-ending': 'gzip, deflate, br',
        'connection': 'keep-alive'
    })
    # session.get("https://stockx.com/api/browse?&_search=EF2367&dataType=product")
    # print(session.cookies)

    row_num = config["start-row"]

    for row in ws.iter_rows(min_row=config["start-row"]): 
        if row[0].value == None:
            break
        print(f'Fetching row {row_num}: {row[0].value}')
        sku = row[0].value
        size = str(row[2].value)

        if sales_db[sku].get(size) is not None: #if fetched before
            row[1].value = productname_db[sku]
            row[8].value = "Stockx"
            row[8].hyperlink = f'https://stockx.com/{productname_db[sku]}'
            sales = sales_db[sku][size]
            print('Skipping, sales exist in database')
            row[4].value = round((sales["last"] * rate), 2)
            row[5].value = round((sales["average"] * rate), 2)
            row[6].value = round((sales["highest_bid"] * rate), 2)
            row[7].value = round((sales["lowest_ask"] * rate), 2)
            now = datetime.now()
            dt_string = now.strftime("%d/%m/%Y %I:%M:%S %p")
            row[10].value = dt_string
            row_num += 1
        else:
            urlkey = search_product(sku, session)
            if urlkey is None:
                break
            row[9].value = "Stockx"
            row[9].hyperlink = f'https://stockx.com/{urlkey}'
            result = product_info(urlkey, size, session)
            if result is None:
                break
            row[1].value = result["title"]
            productname_db[sku] = row[1].value
            if result["uuid"] is None:
                print("Size not found, check if W or Y is needed.")
                break
            sales = get_sales(result["uuid"], sku, size, session)
            if sales is None:
                break
            sales["highest_bid"] = result["highest_bid"]
            sales["lowest_ask"] = result["lowest_ask"]
            sales_db[sku][size] = sales
            row[4].value = round((sales["last"] * rate), 2)
            row[5].value = round((sales["average"] * rate), 2)
            row[6].value = round((sales["highest_bid"] * rate), 2)
            row[7].value = round((sales["lowest_ask"] * rate), 2)
            now = datetime.now()
            dt_string = now.strftime("%d/%m/%Y %I:%M:%S %p")
            row[10].value = dt_string
            row_num += 1
        
    wb.save(resource_path("./stock_book_output.xlsx"))

def search_product(sku, session):
    url = f'https://stockx.com/api/browse?&_search={sku}&dataType=product'
    
    req = session.get(url)
    while req.status_code != 200:
        print(f'Error {req.status_code} at search_product')
        input("Please solve captcha at https://www.stockx.com on an incognito window and press enter")
        req = session.get(url)
    else:
        data = req.json()
        urlkey = data["Products"][0]["urlKey"]
        return urlkey
    
def product_info(urlkey, size, session):
    url = f'https://stockx.com/api/products/{urlkey}?includes=market&currency=USD'

    req = session.get(url)
    while req.status_code != 200:
        print(f'Error {req.status_code} at product_info')
        input("Please solve captcha at https://www.stockx.com on an incognito window and press enter")
        req = session.get(url)
    else:
        data = req.json()
        result = {}
        result["title"] = data["Product"]["title"]
        result["uuid"] = None
        for key in data["Product"]["children"]:
            if (data["Product"]["children"][key]["shoeSize"]) == size:
                uuid = data["Product"]["children"][key]["uuid"]
                highest_bid = data["Product"]["children"][key]["market"]["highestBid"]
                lowest_ask = data["Product"]["children"][key]["market"]["lowestAsk"]
                result["uuid"] = uuid
                result["highest_bid"] = highest_bid
                result["lowest_ask"] = lowest_ask
                break
        return result

def get_sales(uuid, sku, size, session):
    url = f'https://stockx.com/api/products/{uuid}/activity?state=480&currency=USD&limit=3&page=1&sort=createdAt&order=DESC'

    req = session.get(url)
    while req.status_code != 200:
        print(f'Error {req.status_code} at get_sales')
        input("Please solve captcha at https://www.stockx.com on an incognito window and press enter")
        req = session.get(url)
    else:
        data = req.json()
        amount = 0
        limit = 0
        sales = {}
        for i in data["ProductActivity"]:
            if limit == 0:
                sales["last"] = i["localAmount"]
            amount += i["localAmount"]
            limit += 1
        average = amount / limit
        sales["average"] = round(average, 2)
        return sales
    
if __name__ == '__main__':
    main()